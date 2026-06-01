"""
engine_core.py
--------------
Deterministic data engine for the Excel -> Power BI Granularizer.

Responsibilities:
  * load a workbook into plain grids (values AND formulas) for .xlsx/.xls/.csv
  * heuristically guess the layout (header band, id columns, total/banner rows)
  * deterministically reshape a sheet into a tidy / long granular table given a config
  * reconcile the output against the source (no values invented or lost)
  * export to CSV and to a two-sheet XLSX (clean data + transformation log)

Nothing here invents numbers. Detection only *guesses* structure; the actual
reshape is fully determined by the config the user confirms in the UI.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TOTAL_RE = re.compile(r"\b(sub)?total\b|grand\s*total", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #
@dataclass
class Sheet:
    name: str
    values: list[list[Any]]          # cached/computed values (numbers, text)
    formulas: list[list[Any]] | None # formula strings (=...) or None if N/A
    merged: list[tuple[int, int, int, int]] = field(default_factory=list)
    # merged ranges as (min_row, min_col, max_row, max_col), 1-indexed


@dataclass
class Book:
    sheets: dict[str, Sheet]

    @property
    def sheet_names(self) -> list[str]:
        return list(self.sheets.keys())


def _grid_from_ws(ws) -> list[list[Any]]:
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row))
    # normalise ragged rows to equal width
    width = max((len(r) for r in grid), default=0)
    for r in grid:
        r.extend([None] * (width - len(r)))
    return grid


def load_book(file_bytes: bytes, filename: str) -> Book:
    """Load any supported file into a Book of Sheets.

    For .xlsx we read twice with openpyxl: once for cached values, once for
    formula text. For .xls / .csv we only get values (no formula tracing).
    """
    ext = filename.lower().rsplit(".", 1)[-1]

    if ext == "xlsx":
        wb_v = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
        wb_f = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
        sheets: dict[str, Sheet] = {}
        for name in wb_v.sheetnames:
            ws_v = wb_v[name]
            ws_f = wb_f[name]
            merged = [
                (m.min_row, m.min_col, m.max_row, m.max_col)
                for m in ws_v.merged_cells.ranges
            ]
            sheets[name] = Sheet(
                name=name,
                values=_grid_from_ws(ws_v),
                formulas=_grid_from_ws(ws_f),
                merged=merged,
            )
        return Book(sheets)

    if ext == "xls":
        xls = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)
        return Book({n: Sheet(n, df.values.tolist(), None) for n, df in xls.items()})

    if ext == "csv":
        df = pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=object)
        return Book({"CSV": Sheet("CSV", df.values.tolist(), None)})

    raise ValueError(f"Unsupported file type: .{ext}")


def apply_merges(sheet: Sheet) -> list[list[Any]]:
    """Return a copy of the value grid with merged cells filled (top-left value
    propagated across the merged range). Helps header/banner detection."""
    grid = [row[:] for row in sheet.values]
    for (r0, c0, r1, c1) in sheet.merged:
        val = grid[r0 - 1][c0 - 1] if r0 - 1 < len(grid) else None
        for r in range(r0 - 1, min(r1, len(grid))):
            for c in range(c0 - 1, min(c1, len(grid[r]))):
                if grid[r][c] in (None, ""):
                    grid[r][c] = val
    return grid


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _nonempty_count(row: list[Any]) -> int:
    return sum(0 if _is_blank(v) else 1 for v in row)


def to_number(v: Any):
    """Best-effort numeric coercion. Returns float or None (never raises)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    pct = s.endswith("%")
    s = s.replace(",", "").replace("$", "").replace("€", "").replace("£", "").rstrip("%").strip()
    s = s.replace("(", "-").replace(")", "")
    try:
        n = float(s)
        return n / 100.0 if pct else n
    except ValueError:
        return None


def _looks_numeric_row(row: list[Any], min_frac: float = 0.4) -> bool:
    cells = [v for v in row if not _is_blank(v)]
    if not cells:
        return False
    nums = sum(1 for v in cells if to_number(v) is not None)
    return nums / len(cells) >= min_frac


# --------------------------------------------------------------------------- #
# Heuristic detection (only guesses; user confirms in the UI)
# --------------------------------------------------------------------------- #
@dataclass
class LayoutGuess:
    header_start: int          # 0-indexed row where the header band begins
    header_rows: int           # number of header rows
    data_start: int            # 0-indexed first data row
    id_columns: list[int]      # column indexes that look like identifiers
    total_rows: list[int]      # 0-indexed rows that look like totals/subtotals
    width: int


def detect_layout(grid: list[list[Any]]) -> LayoutGuess:
    width = max((len(r) for r in grid), default=0)
    n = len(grid)

    # 1) data start: the first row of the main numeric block. Scan for the first
    #    row that looks numeric AND is followed by another numeric row (avoids
    #    latching onto a stray number in a title/metadata band).
    data_start = None
    for i in range(n):
        if _looks_numeric_row(grid[i]):
            nxt = _looks_numeric_row(grid[i + 1]) if i + 1 < n else False
            if nxt or i == n - 1:
                data_start = i
                break
    if data_start is None:
        # no clear numeric block: fall back to first dense row + 1
        dense_threshold = max(2, int(0.4 * width))
        first_dense = next((i for i, r in enumerate(grid)
                            if _nonempty_count(r) >= dense_threshold), 0)
        data_start = min(first_dense + 1, n - 1)

    # 2) header band: the contiguous run of non-blank text rows immediately above
    #    data_start (handles wide multi-row headers and leading blank rows).
    header_start = data_start
    k = data_start - 1
    while k >= 0 and _nonempty_count(grid[k]) > 0 and not _looks_numeric_row(grid[k]):
        header_start = k
        k -= 1
    if header_start == data_start:          # no text header found above data
        header_start = max(0, data_start - 1)
    header_rows = max(1, data_start - header_start)

    # 3) id columns: leading columns whose data cells are mostly non-numeric
    id_columns = []
    for c in range(width):
        col_cells = [grid[r][c] for r in range(data_start, n) if c < len(grid[r])]
        col_cells = [v for v in col_cells if not _is_blank(v)]
        if not col_cells:
            continue
        non_num = sum(1 for v in col_cells if to_number(v) is None)
        if non_num / len(col_cells) >= 0.6:
            id_columns.append(c)
        else:
            break  # stop at the first numeric (measure) column
    if not id_columns:
        id_columns = [0]
    # never emit a column index outside the actual width
    id_columns = [c for c in id_columns if 0 <= c < width] or [0]

    # 4) total / subtotal rows (match in any id column)
    total_rows = []
    for r in range(data_start, n):
        joined = " ".join(str(grid[r][c]) for c in id_columns if c < len(grid[r]) and not _is_blank(grid[r][c]))
        if TOTAL_RE.search(joined):
            total_rows.append(r)

    return LayoutGuess(header_start, header_rows, data_start, id_columns, total_rows, width)


# --------------------------------------------------------------------------- #
# Reshape (deterministic, driven by confirmed config)
# --------------------------------------------------------------------------- #
@dataclass
class ReshapeConfig:
    header_start: int
    header_rows: int
    data_start: int
    id_columns: list[int]
    header_sep: str = " | "
    drop_totals: bool = True
    use_row_hierarchy: bool = False   # treat "only first col filled" rows as section banners
    hierarchy_name: str = "Section"
    split_variable: bool = False      # split the melted Variable by header_sep
    split_names: list[str] = field(default_factory=list)
    pivot_metric: bool = False        # pivot the LAST split level back into columns
    skip_repeated_headers: bool = True  # drop in-table rows that repeat the header labels


def _flatten_headers(grid, cfg: ReshapeConfig) -> list[str]:
    width = max((len(r) for r in grid), default=0)
    band = [grid[cfg.header_start + k] if cfg.header_start + k < len(grid) else []
            for k in range(cfg.header_rows)]
    # forward-fill each header row horizontally (handles merged/blank header cells)
    filled = []
    for row in band:
        out, last = [], None
        for c in range(width):
            v = row[c] if c < len(row) else None
            if not _is_blank(v):
                last = str(v).strip()
            out.append(last)
        filled.append(out)
    names = []
    for c in range(width):
        parts = [filled[k][c] for k in range(len(filled)) if filled[k][c]]
        # de-dup consecutive identical parts
        dedup = []
        for p in parts:
            if not dedup or dedup[-1] != p:
                dedup.append(p)
        names.append(cfg.header_sep.join(dedup) if dedup else f"col_{get_column_letter(c+1)}")
    # ensure uniqueness
    seen, uniq = {}, []
    for nm in names:
        if nm in seen:
            seen[nm] += 1
            uniq.append(f"{nm}_{seen[nm]}")
        else:
            seen[nm] = 0
            uniq.append(nm)
    return uniq


def reshape(sheet: Sheet, cfg: ReshapeConfig) -> dict:
    grid = apply_merges(sheet)
    width = max((len(r) for r in grid), default=0)

    # --- guard rails: never crash on empty / odd sheets ----------------------
    if width == 0 or len(grid) == 0:
        raise ValueError("This sheet is empty — pick another sheet.")
    # clamp config to the actual grid so a stale guess can't index out of range
    cfg.header_start = max(0, min(cfg.header_start, len(grid) - 1))
    cfg.data_start = max(cfg.header_start + 1, min(cfg.data_start, len(grid)))
    cfg.header_rows = max(1, min(cfg.header_rows, max(1, cfg.data_start - cfg.header_start)))
    cfg.id_columns = sorted({c for c in cfg.id_columns if 0 <= c < width}) or [0]
    if cfg.data_start >= len(grid):
        raise ValueError("No data rows below the header — check 'Data starts at row'.")
    if not [c for c in range(width) if c not in cfg.id_columns]:
        raise ValueError("Every column is marked as an identifier — leave at least one "
                         "measure column unselected so there is something to unpivot.")

    names = _flatten_headers(grid, cfg)

    log: list[str] = []
    log.append(f"Header band: rows {cfg.header_start+1}-{cfg.header_start+cfg.header_rows} "
               f"({cfg.header_rows} row(s)).")
    log.append(f"Data starts at row {cfg.data_start+1}.")

    id_names = [names[c] for c in cfg.id_columns]
    val_cols = [c for c in range(width) if c not in cfg.id_columns]

    # set of header label strings, to recognise repeated sub-header rows mid-table
    header_label_set = {str(names[c]).strip().lower() for c in range(width)}
    measure_labels = {str(grid[cfg.data_start - 1][c]).strip().lower()
                      for c in val_cols
                      if cfg.data_start - 1 < len(grid) and c < len(grid[cfg.data_start - 1])
                      and not _is_blank(grid[cfg.data_start - 1][c])}

    rows, dropped_blank, dropped_total, dropped_repeat, set_aside_totals = [], 0, 0, 0, []
    current_section = None
    for r in range(cfg.data_start, len(grid)):
        row = grid[r]
        if _nonempty_count(row) == 0:
            dropped_blank += 1
            continue
        first_vals = [row[c] for c in cfg.id_columns if c < len(row)]
        joined = " ".join(str(v) for v in first_vals if not _is_blank(v))

        if cfg.drop_totals and TOTAL_RE.search(joined):
            dropped_total += 1
            set_aside_totals.append([row[c] if c < len(row) else None for c in val_cols])
            continue

        # repeated sub-header row: the measure cells echo the header labels
        if cfg.skip_repeated_headers and measure_labels:
            cell_strs = {str(row[c]).strip().lower() for c in val_cols
                         if c < len(row) and not _is_blank(row[c])}
            if cell_strs and cell_strs.issubset(header_label_set | measure_labels):
                # also require that almost none of the measure cells are numeric
                numeric_here = sum(1 for c in val_cols
                                   if c < len(row) and to_number(row[c]) is not None)
                if numeric_here == 0:
                    dropped_repeat += 1
                    continue

        # section banner: only the first id column filled, rest of row empty
        if cfg.use_row_hierarchy:
            only_first = (not _is_blank(row[cfg.id_columns[0]] if cfg.id_columns[0] < len(row) else None)
                          and _nonempty_count(row) == 1)
            if only_first:
                current_section = str(row[cfg.id_columns[0]]).strip()
                continue

        rec = {names[c]: (row[c] if c < len(row) else None) for c in cfg.id_columns}
        if cfg.use_row_hierarchy:
            rec[cfg.hierarchy_name] = current_section
        for c in val_cols:
            rec[names[c]] = row[c] if c < len(row) else None
        rows.append(rec)

    wide = pd.DataFrame(rows)
    log.append(f"Dropped {dropped_blank} blank, {dropped_total} total/subtotal, "
               f"and {dropped_repeat} repeated-header row(s).")
    if cfg.use_row_hierarchy:
        log.append(f"Filled row hierarchy down into column '{cfg.hierarchy_name}'.")

    id_vars = list(id_names)
    if cfg.use_row_hierarchy:
        id_vars = [cfg.hierarchy_name] + id_vars
    value_vars = [names[c] for c in val_cols]

    tidy = wide.melt(id_vars=id_vars, value_vars=value_vars,
                     var_name="Variable", value_name="Value")
    tidy["Value"] = tidy["Value"].map(to_number)
    tidy = tidy.dropna(subset=["Value"]).reset_index(drop=True)
    log.append(f"Unpivoted {len(value_vars)} measure column(s) into long format "
               f"-> {len(tidy)} granular rows.")

    if cfg.split_variable:
        parts = tidy["Variable"].str.split(re.escape(cfg.header_sep), expand=True)
        ncols = parts.shape[1]
        col_names = (cfg.split_names + [f"Level_{i+1}" for i in range(ncols)])[:ncols]
        for i in range(ncols):
            tidy[col_names[i]] = parts[i].str.strip()
        tidy = tidy.drop(columns=["Variable"])
        log.append(f"Split header into columns: {', '.join(col_names)}.")

        # pivot the LAST split level (e.g. Metric) back out into its own columns
        if cfg.pivot_metric and len(col_names) >= 2:
            metric_col = col_names[-1]
            index_cols = [c for c in tidy.columns if c not in (metric_col, "Value")]
            tidy = (tidy
                    .pivot_table(index=index_cols, columns=metric_col,
                                 values="Value", aggfunc="first")
                    .reset_index())
            tidy.columns.name = None
            log.append(f"Pivoted '{metric_col}' back into measure columns "
                       f"-> {len(tidy)} rows, columns: {', '.join(map(str, tidy.columns))}.")

    return {
        "tidy": tidy,
        "wide": wide,
        "log": log,
        "id_vars": id_vars,
        "value_vars": value_vars,
        "set_aside_totals": set_aside_totals,
    }


# --------------------------------------------------------------------------- #
# Reconciliation
# --------------------------------------------------------------------------- #
def reconcile(result: dict, sheet: Sheet, cfg: ReshapeConfig) -> dict:
    """Global sum check: total of all output values vs total of the source
    measure cells (excluding total rows). Confirms nothing was added or lost."""
    grid = apply_merges(sheet)
    width = max((len(r) for r in grid), default=0)
    val_cols = [c for c in range(width) if c not in cfg.id_columns]

    src_total = 0.0
    for r in range(cfg.data_start, len(grid)):
        row = grid[r]
        joined = " ".join(str(row[c]) for c in cfg.id_columns
                          if c < len(row) and not _is_blank(row[c]))
        if cfg.drop_totals and TOTAL_RE.search(joined):
            continue
        if cfg.use_row_hierarchy and _nonempty_count(row) == 1:
            continue
        for c in val_cols:
            n = to_number(row[c] if c < len(row) else None)
            if n is not None:
                src_total += n

    out_total = float(result["tidy"]["Value"].sum())
    diff = out_total - src_total
    ok = abs(diff) < max(1e-6, abs(src_total) * 1e-9)
    return {"source_total": src_total, "output_total": out_total, "difference": diff, "ok": ok}


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def to_xlsx_bytes(df: pd.DataFrame, log_lines: list[str]) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Granular")
        pd.DataFrame({"Transformation log": log_lines}).to_excel(
            xw, index=False, sheet_name="Transformation_Log")
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Workbook analyzer — classify every sheet
# --------------------------------------------------------------------------- #
def analyze_workbook(book: "Book") -> list[dict]:
    """Classify each sheet so the user can see which are worth granularizing.

    Verdicts:
      * 'empty'            : no data.
      * 'already_granular' : a single flat table, one header row, mostly one row
                             per record already -> just export as-is.
      * 'reshapeable'      : cross-tab / multi-row header / wide repeating blocks
                             -> can be unpivoted into a granular table.
      * 'review'           : has data but structure is unclear -> open and confirm.
    """
    out = []
    for nm in book.sheet_names:
        s = book.sheets[nm]
        grid = apply_merges(s)
        width = max((len(r) for r in grid), default=0)
        nonblank = [r for r in grid if _nonempty_count(r) > 0]
        rows = len(nonblank)

        if width == 0 or rows == 0:
            out.append({"sheet": nm, "verdict": "empty", "rows": 0, "cols": width,
                        "why": "No data on this sheet.", "suggest": "Skip."})
            continue

        guess = detect_layout(grid)
        # signals
        has_multirow_header = guess.header_rows >= 2
        # repeating metric labels across the header => wide cross-tab
        header_vals = [grid[guess.header_start + guess.header_rows - 1][c]
                       for c in range(width)
                       if guess.header_start + guess.header_rows - 1 < len(grid)
                       and c < len(grid[guess.header_start + guess.header_rows - 1])]
        header_strs = [str(v).strip().lower() for v in header_vals if not _is_blank(v)]
        repeats = len(header_strs) - len(set(header_strs))
        has_repeating_blocks = repeats >= 2
        has_merges = len(s.merged) > 0
        leading_blanks = guess.header_start
        data_rows = rows - guess.header_rows

        if has_multirow_header or has_repeating_blocks or (has_merges and leading_blanks):
            verdict = "reshapeable"
            why = []
            if has_multirow_header:
                why.append(f"{guess.header_rows}-row header")
            if has_repeating_blocks:
                why.append("repeating metric columns (cross-tab)")
            if has_merges:
                why.append(f"{len(s.merged)} merged ranges")
            suggest = (f"Header at row {guess.header_start+1}, data at "
                       f"row {guess.data_start+1}; unpivot the repeating columns.")
        elif guess.header_rows == 1 and data_rows >= 3 and not has_repeating_blocks:
            verdict = "already_granular"
            why = ["single header row, flat record layout"]
            suggest = "Likely usable as-is; export without unpivoting."
        else:
            verdict = "review"
            why = ["has data but layout is ambiguous"]
            suggest = "Open this sheet and confirm the header/data rows."

        out.append({"sheet": nm, "verdict": verdict, "rows": data_rows,
                    "cols": width, "why": "; ".join(why) if isinstance(why, list) else why,
                    "suggest": suggest,
                    "header_row": guess.header_start + 1,
                    "data_row": guess.data_start + 1})
    return out
