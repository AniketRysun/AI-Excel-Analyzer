"""
app.py — Excel -> Power BI Granularizer (Streamlit)

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

The app follows the spec's flow: Upload -> auto-detect layout -> you confirm/adjust
-> deterministic reshape -> preview granular table + log + reconciliation -> download.
A second tab traces formulas back to their source cells.
"""
from __future__ import annotations

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter, column_index_from_string

import engine_core as E
import formula_trace as T
import hhw_parser as H
import gtg_parser as G
import gtg_parser_2025 as G25
import bu_parser as BU
import eventlog_parser as EL

import re as _re


def _year_from_filename(name: str) -> str | None:
    """Pull a 4-digit year (2000-2099) from the uploaded filename, if present."""
    m = _re.search(r"(20\d{2})", name or "")
    return m.group(1) if m else None

st.set_page_config(page_title="Excel → Power BI Granularizer", layout="wide")

st.title("Excel → Power BI Granularizer")
st.caption(
    "Upload a messy spreadsheet, confirm the detected structure, and download a tidy, "
    "granular table ready for Power BI. The reshape is fully deterministic — no numbers "
    "are ever invented; anything uncertain is flagged for you to confirm."
)

uploaded = st.file_uploader("Upload a workbook", type=["xlsx", "xls", "csv"])
if not uploaded:
    st.info("Upload an .xlsx, .xls, or .csv file to begin. "
            "No file handy? Run `python make_sample.py` to create one.")
    st.stop()

try:
    book = E.load_book(uploaded.getvalue(), uploaded.name)
except Exception as exc:  # noqa: BLE001
    st.error(f"Could not read the file: {exc}")
    st.stop()

tab_scan, tab_reshape, tab_trace = st.tabs(
    ["Scan workbook", "Reshape to granular table", "Trace formulas"])

# --------------------------------------------------------------------------- #
# Tab 0 — scan every sheet and give a verdict
# --------------------------------------------------------------------------- #
with tab_scan:
    st.write("A verdict for every sheet — which ones can become granular tables, "
             "which are already flat, and which to skip.")
    report = E.analyze_workbook(book)
    for row in report:
        if row["verdict"] != "empty" and H.looks_like_hhw(book.sheets[row["sheet"]]):
            row["verdict"] = "reshapeable (H&HW)"
            row["suggest"] = "Use the Reshape tab — the H&HW parser handles this automatically."
    rep_df = pd.DataFrame(report)[["sheet", "verdict", "rows", "cols", "why", "suggest"]]
    rep_df.columns = ["Sheet", "Verdict", "Data rows", "Cols", "Why", "Suggested action"]
    st.dataframe(rep_df, use_container_width=True, height=560)
    st.caption("Verdicts are heuristic guesses to point you at the right sheets — "
               "confirm in the Reshape tab before trusting any output.")

# --------------------------------------------------------------------------- #
# Tab 1 — reshape
# --------------------------------------------------------------------------- #
with tab_reshape:
    # default to the first sheet that actually has data (skip empty Dashboards etc.)
    def _sheet_has_data(nm: str) -> bool:
        g = book.sheets[nm].values
        return any(E._nonempty_count(r) > 0 for r in g)

    nonempty_sheets = [nm for nm in book.sheet_names if _sheet_has_data(nm)]
    default_idx = book.sheet_names.index(nonempty_sheets[0]) if nonempty_sheets else 0
    sheet_name = st.selectbox("Sheet", book.sheet_names, index=default_idx, key="reshape_sheet")
    sheet = book.sheets[sheet_name]
    grid = E.apply_merges(sheet)
    width = max((len(r) for r in grid), default=0)

    if width == 0 or not any(E._nonempty_count(r) > 0 for r in grid):
        st.warning(f"Sheet **{sheet_name}** is empty. Pick a sheet that contains data.")
        st.stop()

    with st.expander("Raw sheet preview (first 15 rows)", expanded=True):
        prev = pd.DataFrame(
            [[("" if E._is_blank(v) else v) for v in row[:width]] for row in grid[:15]],
            columns=[get_column_letter(c + 1) for c in range(width)],
        )
        st.dataframe(prev, use_container_width=True)

    # ---- dedicated path for H&HW cross-tabs (handles the irregular layout) ----
    _handled = False
    if BU.has_bu_table(sheet):
        _handled = True
        yr = _year_from_filename(uploaded.name)
        st.success("Detected the Global Business Units table — transposing into "
                   "one row per business-unit/region.")
        yr_in = st.text_input("Year (from filename; edit if needed)", yr or "", key="bu_year")
        try:
            bres = BU.parse(sheet, year=yr_in or None)
            bdf = bres["tidy"]
            st.subheader(f"Granular table — {len(bdf)} rows")
            st.dataframe(bdf, use_container_width=True, height=420)
            with st.expander("Transformation log"):
                for line in bres["log"]:
                    st.write("•", line)
            d1, d2 = st.columns(2)
            d1.download_button("Download CSV", E.to_csv_bytes(bdf),
                               file_name=f"{sheet_name}_granular.csv", mime="text/csv")
            d2.download_button("Download XLSX (data + log)",
                               E.to_xlsx_bytes(bdf, bres["log"]),
                               file_name=f"{sheet_name}_granular.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Business Units parse failed: {exc}")
    elif EL.looks_like_eventlog(sheet):
        _handled = True
        yr = _year_from_filename(uploaded.name)
        st.success("Detected an Event / Incident Log — already granular. Keeping every "
                   "column (including any new ones) and adding Year.")
        yr_in = st.text_input("Year (from filename; edit if needed)", yr or "", key="el_year")
        try:
            eres = EL.parse(sheet, year=yr_in or None)
            edf = eres["tidy"]
            st.subheader(f"Granular table — {len(edf)} rows, {len(edf.columns)} columns")
            st.dataframe(edf, use_container_width=True, height=420)
            with st.expander("Transformation log"):
                for line in eres["log"]:
                    st.write("•", line)
            d1, d2 = st.columns(2)
            d1.download_button("Download CSV", E.to_csv_bytes(edf),
                               file_name=f"{sheet_name}_granular.csv", mime="text/csv")
            d2.download_button("Download XLSX (data + log)",
                               E.to_xlsx_bytes(edf, eres["log"]),
                               file_name=f"{sheet_name}_granular.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Event Log parse failed: {exc}")
    elif G25.looks_like_gtg_2025(sheet):
        _handled = True
        yr = _year_from_filename(uploaded.name)
        st.success("Detected a 2025-style GTG sheet (per-business-unit layout) — "
                   "using the dedicated parser.")
        yr_in = st.text_input("Year (from filename; edit if needed)", yr or "", key="gtg25_year")
        try:
            gres = G25.parse(sheet, year=yr_in or None)
            gdf = gres["tidy"]
            st.subheader(f"Granular table — {len(gdf)} rows")
            st.dataframe(gdf, use_container_width=True, height=420)
            with st.expander("Transformation log"):
                for line in gres["log"]:
                    st.write("•", line)
            d1, d2 = st.columns(2)
            d1.download_button("Download CSV", E.to_csv_bytes(gdf),
                               file_name=f"{sheet_name}_granular.csv", mime="text/csv")
            d2.download_button("Download XLSX (data + log)",
                               E.to_xlsx_bytes(gdf, gres["log"]),
                               file_name=f"{sheet_name}_granular.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:  # noqa: BLE001
            st.error(f"2025 GTG parse failed: {exc}")
    elif G.looks_like_gtg(sheet):
        _handled = True
        yr = _year_from_filename(uploaded.name)
        st.success("Detected a Global Technical Governance sheet — using the dedicated "
                   "parser (granular Location rows, skipping the aggregated tables).")
        c_yr1, c_yr2 = st.columns([1, 3])
        yr_in = c_yr1.text_input("Year (from filename; edit if needed)", yr or "")
        try:
            gres = G.parse(sheet, year=yr_in or None)
            gdf = gres["tidy"]
            st.subheader(f"Granular table — {len(gdf)} rows")
            st.dataframe(gdf, use_container_width=True, height=420)
            with st.expander("Transformation log"):
                for line in gres["log"]:
                    st.write("•", line)
            d1, d2 = st.columns(2)
            d1.download_button("Download CSV", E.to_csv_bytes(gdf),
                               file_name=f"{sheet_name}_granular.csv", mime="text/csv")
            d2.download_button("Download XLSX (data + log)",
                               E.to_xlsx_bytes(gdf, gres["log"]),
                               file_name=f"{sheet_name}_granular.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:  # noqa: BLE001
            st.error(f"GTG parse failed: {exc}")
    elif H.looks_like_hhw(sheet):
        _handled = True
        st.success("Detected an H&HW (Headcount & Hours Worked) cross-tab — "
                   "using the dedicated parser.")
        include_ytd = st.checkbox("Include 'Year to Date Average' as a row "
                                  "(off = months only, recommended)", value=False)
        try:
            hres = H.parse(sheet, include_ytd=include_ytd)
            hdf = hres["tidy"]
            st.subheader(f"Granular table — {len(hdf)} rows")
            st.dataframe(hdf, use_container_width=True, height=420)
            with st.expander("Transformation log"):
                for line in hres["log"]:
                    st.write("•", line)
            d1, d2 = st.columns(2)
            d1.download_button("Download CSV", E.to_csv_bytes(hdf),
                               file_name=f"{sheet_name}_granular.csv", mime="text/csv")
            d2.download_button("Download XLSX (data + log)",
                               E.to_xlsx_bytes(hdf, hres["log"]),
                               file_name=f"{sheet_name}_granular.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:  # noqa: BLE001
            st.error(f"H&HW parse failed: {exc}")

    if not _handled:
        flat_mode = st.checkbox(
            "This sheet is already a flat table (one row per record) — just keep all "
            "columns and add Year, don't unpivot", value=False, key="flat_mode")
        if flat_mode:
            yr = _year_from_filename(uploaded.name)
            yr_in = st.text_input("Year (from filename; edit if needed)", yr or "", key="flat_year")
            try:
                fres = EL.parse_flat(sheet, year=yr_in or None)
                fdf = fres["tidy"]
                st.subheader(f"Flat table — {len(fdf)} rows, {len(fdf.columns)} columns")
                st.dataframe(fdf, use_container_width=True, height=420)
                with st.expander("Transformation log"):
                    for line in fres["log"]:
                        st.write("•", line)
                d1, d2 = st.columns(2)
                d1.download_button("Download CSV", E.to_csv_bytes(fdf),
                                   file_name=f"{sheet_name}_flat.csv", mime="text/csv")
                d2.download_button("Download XLSX (data + log)",
                                   E.to_xlsx_bytes(fdf, fres["log"]),
                                   file_name=f"{sheet_name}_flat.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.stop()
            except Exception as exc:  # noqa: BLE001
                st.error(f"Flat-table parse failed: {exc}")
                st.stop()

        guess = E.detect_layout(grid)
        st.markdown("**Detected layout** (adjust anything that looks wrong):")

        n_rows = max(1, len(grid))

        def _clamp(v: int, lo: int, hi: int) -> int:
            return max(lo, min(int(v), hi))

        c1, c2, c3 = st.columns(3)
        header_start = c1.number_input("Header starts at row", 1, n_rows,
                                       _clamp(guess.header_start + 1, 1, n_rows))
        header_rows = c2.number_input("Number of header rows", 1, 10,
                                      _clamp(guess.header_rows, 1, 10))
        data_start = c3.number_input("Data starts at row", 1, n_rows,
                                     _clamp(guess.data_start + 1, 1, n_rows))

        letters = [get_column_letter(c + 1) for c in range(width)]
        # keep only detected id columns that are within the actual column range
        default_ids = [get_column_letter(c + 1) for c in guess.id_columns
                       if 0 <= c < width]
        default_ids = [l for l in default_ids if l in letters]
        if not default_ids and letters:
            default_ids = [letters[0]]
        id_letters = st.multiselect("Identifier columns (kept as-is; everything else is unpivoted)",
                                    letters, default=default_ids)

        c4, c5 = st.columns(2)
        use_hier = c4.checkbox("Treat single-cell rows as section banners (fill down)", value=True)
        hier_name = c4.text_input("Section column name", "Section", disabled=not use_hier)
        drop_totals = c5.checkbox("Drop total / subtotal rows", value=True)

        c6, c7 = st.columns(2)
        sep = c6.text_input("Header level separator", " | ")
        split_var = c7.checkbox("Split the unpivoted header into multiple columns", value=True)
        split_names_raw = c7.text_input("Names for split columns (comma-separated)",
                                        "Month, Metric", disabled=not split_var)

        cfg = E.ReshapeConfig(
            header_start=int(header_start) - 1,
            header_rows=int(header_rows),
            data_start=int(data_start) - 1,
            id_columns=[column_index_from_string(l) - 1 for l in id_letters] or [0],
            header_sep=sep,
            drop_totals=drop_totals,
            use_row_hierarchy=use_hier,
            hierarchy_name=hier_name or "Section",
            split_variable=split_var,
            split_names=[s.strip() for s in split_names_raw.split(",") if s.strip()],
        )

        try:
            res = E.reshape(sheet, cfg)
            rec = E.reconcile(res, sheet, cfg)
        except Exception as exc:  # noqa: BLE001
            st.error(f"Reshape failed with the current settings: {exc}")
            st.stop()

        st.subheader(f"Granular table — {len(res['tidy'])} rows")
        st.dataframe(res["tidy"], use_container_width=True, height=380)

        m1, m2, m3 = st.columns(3)
        m1.metric("Source total", f"{rec['source_total']:,.2f}")
        m2.metric("Output total", f"{rec['output_total']:,.2f}")
        m3.metric("Reconciles?", "✅ yes" if rec["ok"] else f"⚠️ off by {rec['difference']:,.2f}")

        with st.expander("Transformation log"):
            for line in res["log"]:
                st.write("•", line)

        d1, d2 = st.columns(2)
        d1.download_button("Download CSV", E.to_csv_bytes(res["tidy"]),
                           file_name=f"{sheet_name}_granular.csv", mime="text/csv")
        d2.download_button("Download XLSX (data + log)",
                           E.to_xlsx_bytes(res["tidy"], res["log"]),
                           file_name=f"{sheet_name}_granular.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # --------------------------------------------------------------------------- #
    # Tab 2 — formula tracing
    # --------------------------------------------------------------------------- #
with tab_trace:
    if sheet.formulas is None:
        st.warning("Formula tracing needs an .xlsx file (no formula data in .xls / .csv).")
    else:
        tsheet = st.selectbox("Sheet", book.sheet_names, key="trace_sheet")
        mode = st.radio("Mode", ["Trace one cell", "Scan whole sheet"], horizontal=True)

        if mode == "Trace one cell":
            coord = st.text_input("Cell (e.g. M1)", "M1").strip().upper()
            if st.button("Trace") and coord:
                r = T.trace_cell(book, tsheet, coord)
                st.write(f"**Formula:** `{r['formula']}`")
                badge = {"traceable": "✅", "untraceable": "⚠️", "not_a_formula": "ℹ️"}.get(r["status"], "")
                st.write(f"**Status:** {badge} {r['status']} — {r['reason']}")
                if r.get("lineage"):
                    st.write(f"**Lineage:** `{r['lineage']}`")
                for s in r.get("sources", []):
                    st.write(f"Source `{s['sheet']}!{s['range']}`"
                             + (f" — {s['note']}" if s.get("note") else ""))
                    if s.get("values"):
                        st.dataframe(pd.DataFrame(s["values"]), use_container_width=True)
        else:
            if st.button("Scan sheet for formulas"):
                rows = T.trace_sheet(book, tsheet)
                if not rows:
                    st.info("No formulas found on this sheet.")
                else:
                    st.dataframe(pd.DataFrame(
                        [{"Cell": x["cell"], "Status": x["status"],
                          "Formula": x["formula"], "Detail": x["reason"]} for x in rows]),
                        use_container_width=True, height=400)
