"""
make_sample.py
--------------
Generates `sample_workbook.xlsx` with:
  * 'Report'     : a messy cross-tab (title rows, region banners, country rows,
                   multi-row Quarter->Month->Metric headers, subtotal rows),
                   with some formulas (HrsWorked = Headcount * WorkWeek).
  * 'Timesheets' : a flat detail tab.
  * a cell on 'Report' that references Timesheets via =SUM(...) to exercise
    cross-sheet formula tracing.

Run:  python make_sample.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = "Report"

# --- title / metadata band ---
ws["A1"] = "2026 Headcount & Hours Worked - Business Assurance Globally"
ws["A1"].font = Font(bold=True)
ws["A2"] = "Annual: 2080 + 20% OT per Person FULL TIME"

# --- multi-row header band (rows 4-5) ---
# row 4: quarter/month level ; row 5: metric level
ws["A4"] = None
ws["B4"] = "Quarter 1"; ws["E4"] = "Quarter 1"; ws["H4"] = "Quarter 2"
ws["B4"] = "January"; ws["E4"] = "February"; ws["H4"] = "March"
ws["A5"] = "Country"
ws["B5"] = "Headcount"; ws["C5"] = "Work Week"; ws["D5"] = "Hrs Worked"
ws["E5"] = "Headcount"; ws["F5"] = "Work Week"; ws["G5"] = "Hrs Worked"
ws["H5"] = "Headcount"; ws["I5"] = "Work Week"; ws["J5"] = "Hrs Worked"
# fix month placement (B/E/H are month group starts)
ws["B4"] = "January"; ws["C4"] = "January"; ws["D4"] = "January"
ws["E4"] = "February"; ws["F4"] = "February"; ws["G4"] = "February"
ws["H4"] = "March"; ws["I4"] = "March"; ws["J4"] = "March"

row = 6


def country(name, hc1, ww1, hc2, ww2, hc3, ww3):
    global row
    ws.cell(row=row, column=1, value=name)
    for i, (hc, ww) in enumerate([(hc1, ww1), (hc2, ww2), (hc3, ww3)]):
        base = 2 + i * 3
        ws.cell(row=row, column=base, value=hc)
        ws.cell(row=row, column=base + 1, value=ww)
        # Hrs Worked stored as a value (Headcount * Work Week)
        ws.cell(row=row, column=base + 2, value=hc * ww)
    row += 1


def banner(text):
    global row
    ws.cell(row=row, column=1, value=text)
    row += 1


def subtotal(label, first_row, last_row):
    global row
    ws.cell(row=row, column=1, value=label)
    for col in range(2, 11):
        L = ws.cell(row=first_row, column=col).coordinate
        U = ws.cell(row=last_row, column=col).coordinate
        ws.cell(row=row, column=col, value=f"=SUM({L}:{U})")
    row += 1


banner("Business Assurance - EMEA")
start = row
country("France", 1, 192, 1, 192, 1, 240)
country("Spain", 10, 192, 10, 192, 10, 240)
subtotal("TOTAL HEADCOUNT AND HOURS WORKED", start, row - 1)
row += 1  # blank separator

banner("Business Assurance - LATAM")
start = row
country("Argentina", 2, 192, 2, 192, 2, 240)
country("Brazil", 2, 192, 2, 192, 2, 240)
country("Mexico", 1, 192, 1, 192, 1, 240)
subtotal("TOTAL HEADCOUNT AND HOURS WORKED", start, row - 1)

# --- a cross-sheet formula to trace ---
ws["L1"] = "Total timesheet hours (cross-sheet):"
ws["M1"] = "=SUM(Timesheets!C2:C6)"
ws["L2"] = "Volatile example (untraceable):"
ws["M2"] = "=OFFSET(B7,0,0)"

# --- Timesheets detail tab ---
ts = wb.create_sheet("Timesheets")
ts.append(["Employee", "Country", "Hours"])
for emp, ctry, hrs in [("E1", "France", 38), ("E2", "Spain", 40),
                       ("E3", "Brazil", 42), ("E4", "Mexico", 37), ("E5", "Spain", 41)]:
    ts.append([emp, ctry, hrs])

wb.save("sample_workbook.xlsx")
print("wrote sample_workbook.xlsx")
